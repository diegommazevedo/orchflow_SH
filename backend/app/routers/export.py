"""
export.py — Exportação de dados em PDF e XLSX.

Endpoints:
  GET /api/export/backlog/{project_id}?format=pdf|xlsx
  GET /api/export/project/{project_id}/summary?format=pdf
  GET /api/export/time/{project_id}?format=xlsx

Leis respeitadas:
  - Somente leitura — zero escrita no banco
  - Nenhum wizard necessário (dados já foram confirmados)
  - CORS mantido (declarado no main.py)
"""
from __future__ import annotations

import io
import json
import uuid
from datetime import date
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.project import Project
from app.models.task import Task, EisenhowerQuadrant as EQ

router = APIRouter()

# ── Paleta de quadrantes ──────────────────────────────────────────────────────

Q_COLORS_HEX = {
    "q1": "f43f5e",
    "q2": "f59e0b",
    "q3": "48cae4",
    "q4": "9c9cb8",
}
Q_LABELS = {"q1": "Q1 — Urgente + Importante", "q2": "Q2 — Importante",
            "q3": "Q3 — Urgente", "q4": "Q4 — Descarta"}
STATUS_LABELS = {"backlog": "Backlog", "in_progress": "Em andamento",
                 "blocked": "Impedimento", "done": "Concluído"}

QUADRANT_ORDER = ["q1", "q2", "q3", "q4"]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_name(name: str, maxlen: int = 30) -> str:
    """Nome seguro para usar em filename."""
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in name)[:maxlen]

def _parse_project_id(raw: str) -> uuid.UUID:
    try:
        return uuid.UUID(raw)
    except ValueError:
        raise HTTPException(status_code=422, detail="project_id inválido")

def _get_project(project_id: str, db: Session) -> Project:
    pid = _parse_project_id(project_id)
    p = db.query(Project).filter(Project.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Projeto não encontrado")
    return p

def _get_tasks(project_id: str, db: Session) -> list[Task]:
    pid = _parse_project_id(project_id)
    return db.query(Task).filter(Task.project_id == pid).order_by(Task.quadrant, Task.created_at).all()

def _time_label(mins: int) -> str:
    h, m = divmod(mins, 60)
    return f"{h}:{m:02d}" if h else f"0:{m:02d}"

def _parse_meta(project: Project) -> dict[str, Any]:
    """Tenta extrair metadados do JSON armazenado em project.description."""
    try:
        return json.loads(project.description or "{}")
    except (json.JSONDecodeError, TypeError):
        return {}


# ══ XLSX helpers ══════════════════════════════════════════════════════════════

def _build_backlog_xlsx(project: Project, tasks: list[Task]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    ACCENT   = "7C6DF0"
    WHITE    = "FFFFFF"
    LIGHT_BG = "F4F4F8"
    BORDER_C = "CCCCDD"

    thin_side = Side(style="thin", color=BORDER_C)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"

    # ── Título ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Backlog — {project.name}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=ACCENT)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    date_cell = ws["A2"]
    date_cell.value = f"Gerado em {date.today().strftime('%d/%m/%Y')}  ·  {len(tasks)} tasks"
    date_cell.font = Font(name="Calibri", italic=True, size=10, color="888899")
    date_cell.fill = PatternFill("solid", fgColor="EEEEF4")
    date_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── Cabeçalho ──────────────────────────────────────────────────────────────
    headers = ["Título", "Quadrante", "Status", "Responsável", "Prazo", "Tempo (min)", "Descrição"]
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_i, value=h)
        cell.font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor="4A4A7A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 20

    # ── Dados ──────────────────────────────────────────────────────────────────
    for row_i, task in enumerate(tasks, start=4):
        q_hex = Q_COLORS_HEX.get(str(task.quadrant).replace("EisenhowerQuadrant.", ""), "9C9CB8")
        row_bg = LIGHT_BG if row_i % 2 == 0 else WHITE
        values = [
            task.title,
            Q_LABELS.get(str(task.quadrant), str(task.quadrant)),
            STATUS_LABELS.get(str(task.status), str(task.status)),
            "",  # assignee_hint não é armazenado diretamente
            task.due_date_iso or "",
            task.time_spent_minutes,
            task.description or "",
        ]
        for col_i, val in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_i == 7))
            cell.fill = PatternFill("solid", fgColor=row_bg)
            # Coluna Quadrante: colorir com cor do quadrante
            if col_i == 2:
                cell.font = Font(name="Calibri", size=10, bold=True, color=q_hex)

    # ── Larguras ───────────────────────────────────────────────────────────────
    col_widths = [48, 28, 18, 18, 14, 14, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Linha de totais ────────────────────────────────────────────────────────
    total_row = len(tasks) + 4
    total_time = sum(t.time_spent_minutes for t in tasks)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=total_row, column=6, value=total_time).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=6).fill = PatternFill("solid", fgColor=ACCENT)
    for c in range(1, 8):
        ws.cell(row=total_row, column=c).border = thin_border

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_time_xlsx(project: Project, tasks: list[Task]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ACCENT  = "7C6DF0"
    WHITE   = "FFFFFF"
    LIGHT   = "F4F4F8"
    BORDER_C = "CCCCDD"
    thin_side = Side(style="thin", color=BORDER_C)
    thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tempo"

    # Título
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"Relatório de Tempo — {project.name}"
    t.font = Font(bold=True, size=13, color=WHITE)
    t.fill = PatternFill("solid", fgColor=ACCENT)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    # Cabeçalho
    headers = ["Título", "Quadrante", "Status", "Tempo (min)", "Tempo (h:mm)", "% do total"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font  = Font(bold=True, size=10, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor="4A4A7A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[2].height = 18

    total_time = sum(t.time_spent_minutes for t in tasks) or 1

    for r, task in enumerate(tasks, start=3):
        pct = round(task.time_spent_minutes / total_time * 100, 1)
        bg = LIGHT if r % 2 == 0 else WHITE
        row_vals = [
            task.title,
            Q_LABELS.get(str(task.quadrant), str(task.quadrant)),
            STATUS_LABELS.get(str(task.status), str(task.status)),
            task.time_spent_minutes,
            _time_label(task.time_spent_minutes),
            pct,
        ]
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font   = Font(size=10)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            if c == 6:  # %
                cell.number_format = "0.0%"
                cell.value = pct / 100

    # Total
    tr = len(tasks) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=tr, column=1).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=tr, column=4, value=sum(t.time_spent_minutes for t in tasks)).font = Font(bold=True, color=WHITE)
    ws.cell(row=tr, column=4).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=tr, column=5, value=_time_label(sum(t.time_spent_minutes for t in tasks))).font = Font(bold=True, color=WHITE)
    ws.cell(row=tr, column=5).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=tr, column=6, value=1.0).font = Font(bold=True, color=WHITE)
    ws.cell(row=tr, column=6).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(row=tr, column=6).number_format = "0.0%"
    for c in range(1, 7):
        ws.cell(row=tr, column=c).border = thin

    for i, w in enumerate([48, 28, 18, 14, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══ PDF helpers ═══════════════════════════════════════════════════════════════

def _hex_color(hex_str: str):
    from reportlab.lib.colors import HexColor
    return HexColor(f"#{hex_str}")

def _build_backlog_pdf(project: Project, tasks: list[Task]) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=2*cm,
    )

    ACCENT  = _hex_color("7C6DF0")
    WHITE   = colors.white
    LIGHT   = _hex_color("F4F4F8")
    DARK    = _hex_color("1a1a2e")
    MUTED   = _hex_color("888899")

    styles = {
        "title": ParagraphStyle("title", fontSize=18, fontName="Helvetica-Bold",
                                textColor=WHITE, leading=22),
        "sub":   ParagraphStyle("sub", fontSize=9, fontName="Helvetica",
                                textColor=MUTED, leading=13),
        "qhead": ParagraphStyle("qhead", fontSize=11, fontName="Helvetica-Bold",
                                textColor=DARK, spaceBefore=10, spaceAfter=4),
        "cell":  ParagraphStyle("cell", fontSize=8, fontName="Helvetica",
                                textColor=DARK, leading=11),
        "footer": ParagraphStyle("footer", fontSize=8, fontName="Helvetica",
                                 textColor=MUTED, alignment=1),
    }

    story = []

    # ── Cabeçalho ──────────────────────────────────────────────────────────────
    header_data = [[
        Paragraph(f"Backlog — {project.name}", styles["title"]),
        Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}", styles["sub"]),
    ]]
    header_table = Table(header_data, colWidths=["70%", "30%"])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), ACCENT),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 12),
        ("RIGHTPADDING", (0,0), (-1,-1), 12),
        ("TOPPADDING",   (0,0), (-1,-1), 10),
        ("BOTTOMPADDING",(0,0), (-1,-1), 10),
        ("ALIGN",     (1,0), (1,0), "RIGHT"),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 8))

    total_time = sum(t.time_spent_minutes for t in tasks)
    done_count = sum(1 for t in tasks if str(t.status) in ("done", "TaskStatus.done"))
    summary_data = [[
        f"{len(tasks)} tasks",
        f"{done_count} concluídas",
        f"{len(tasks) - done_count} pendentes",
        f"Tempo total: {_time_label(total_time)}",
    ]]
    summary_table = Table(summary_data, colWidths=["25%","25%","25%","25%"])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), LIGHT),
        ("FONTSIZE",     (0,0), (-1,-1), 9),
        ("FONTNAME",     (0,0), (-1,-1), "Helvetica"),
        ("TEXTCOLOR",    (0,0), (-1,-1), MUTED),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("BOX", (0,0), (-1,-1), 0.5, _hex_color("DDDDEE")),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 12))

    # ── Tasks por quadrante ────────────────────────────────────────────────────
    col_w = [9*cm, 2.8*cm, 2.8*cm, 1.8*cm, 1.8*cm]
    col_hdrs = ["Título", "Quadrante", "Status", "Prazo", "Tempo"]

    for q in QUADRANT_ORDER:
        q_tasks = [t for t in tasks
                   if str(t.quadrant).replace("EisenhowerQuadrant.", "") == q]
        if not q_tasks:
            continue

        q_hex = Q_COLORS_HEX[q]
        q_color = _hex_color(q_hex)

        # Título do quadrante
        story.append(Paragraph(Q_LABELS[q], ParagraphStyle(
            "qh", fontSize=10, fontName="Helvetica-Bold",
            textColor=q_color, spaceBefore=8, spaceAfter=3,
        )))

        data = [col_hdrs]
        for task in q_tasks:
            data.append([
                Paragraph(task.title, styles["cell"]),
                Q_LABELS[q].split(" — ")[0],
                STATUS_LABELS.get(str(task.status).replace("TaskStatus.", ""), str(task.status)),
                task.due_date_iso or "—",
                _time_label(task.time_spent_minutes),
            ])

        tbl = Table(data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            # Cabeçalho
            ("BACKGROUND",    (0,0), (-1,0), _hex_color("4A4A7A")),
            ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 8),
            ("ALIGN",         (0,0), (-1,0), "CENTER"),
            ("TOPPADDING",    (0,0), (-1,0), 5),
            ("BOTTOMPADDING", (0,0), (-1,0), 5),
            # Dados
            ("FONTSIZE",  (0,1), (-1,-1), 8),
            ("FONTNAME",  (0,1), (-1,-1), "Helvetica"),
            ("TEXTCOLOR", (0,1), (-1,-1), DARK),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [WHITE, LIGHT]),
            ("TOPPADDING",    (0,1), (-1,-1), 4),
            ("BOTTOMPADDING", (0,1), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ("GRID",      (0,0), (-1,-1), 0.4, _hex_color("DDDDEE")),
            # Borda esquerda colorida
            ("LINEBEFOREEACH", (0,1), (0,-1), 2.5, q_color),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 6))

    # ── Rodapé ─────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", color=_hex_color("DDDDEE"), thickness=0.5))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"OrchFlow — {project.name} · {len(tasks)} tasks · Tempo total: {_time_label(total_time)} · {date.today().strftime('%d/%m/%Y')}",
        styles["footer"],
    ))

    doc.build(story)
    buf.seek(0)
    return buf


def _build_summary_pdf(project: Project, tasks: list[Task]) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=1.5*cm, bottomMargin=2*cm,
    )

    ACCENT = _hex_color("7C6DF0")
    WHITE  = colors.white
    LIGHT  = _hex_color("F4F4F8")
    DARK   = _hex_color("1a1a2e")
    MUTED  = _hex_color("888899")

    meta = _parse_meta(project)
    client  = meta.get("client") or ""
    start   = meta.get("start_date") or ""
    end     = meta.get("end_date") or ""
    summary = meta.get("contract_summary") or str(project.description or "")

    total  = len(tasks)
    done   = sum(1 for t in tasks if str(t.status) in ("done", "TaskStatus.done"))
    in_prog = [t for t in tasks if str(t.status) in ("in_progress", "TaskStatus.in_progress")]
    pct_done = int(done / total * 100) if total else 0

    story = []

    # Cabeçalho
    h_data = [[
        Paragraph(project.name, ParagraphStyle("n", fontSize=20, fontName="Helvetica-Bold",
                                               textColor=WHITE, leading=24)),
        Paragraph(f"Resumo do Projeto<br/>{date.today().strftime('%d/%m/%Y')}",
                  ParagraphStyle("d", fontSize=9, fontName="Helvetica",
                                 textColor=_hex_color("CCCCEE"), leading=13, alignment=2)),
    ]]
    ht = Table(h_data, colWidths=["65%","35%"])
    ht.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), ACCENT),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0),(-1,-1), 14),
        ("BOTTOMPADDING",(0,0),(-1,-1), 14),
        ("LEFTPADDING",  (0,0),(0,0),  14),
        ("RIGHTPADDING", (1,0),(1,0),  14),
    ]))
    story.append(ht)
    story.append(Spacer(1, 12))

    # Metadados do projeto
    if client or start or end:
        meta_rows = []
        if client:  meta_rows.append(["Cliente", client])
        if start:   meta_rows.append(["Início", start])
        if end:     meta_rows.append(["Fim", end])
        mt = Table(meta_rows, colWidths=[3*cm, 10*cm])
        mt.setStyle(TableStyle([
            ("FONTSIZE",  (0,0),(-1,-1), 9),
            ("FONTNAME",  (0,0),(0,-1), "Helvetica-Bold"),
            ("FONTNAME",  (1,0),(1,-1), "Helvetica"),
            ("TEXTCOLOR", (0,0),(-1,-1), DARK),
            ("TOPPADDING", (0,0),(-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ]))
        story.append(mt)
        story.append(Spacer(1, 8))

    if summary and len(summary) < 500:
        story.append(Paragraph(summary, ParagraphStyle("desc", fontSize=9,
                                                        fontName="Helvetica",
                                                        textColor=MUTED, leading=13)))
        story.append(Spacer(1, 10))

    # Progresso
    story.append(Paragraph("Progresso",
                            ParagraphStyle("sh", fontSize=11, fontName="Helvetica-Bold",
                                           textColor=DARK, spaceBefore=4, spaceAfter=6)))

    prog_data = [["Tasks totais", "Concluídas", "Pendentes", "Progresso"]]
    prog_data.append([str(total), str(done), str(total - done), f"{pct_done}%"])
    pt = Table(prog_data, colWidths=["25%","25%","25%","25%"])
    pt.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,0), _hex_color("4A4A7A")),
        ("TEXTCOLOR",    (0,0),(-1,0), WHITE),
        ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0),(-1,-1), 10),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [LIGHT]),
        ("GRID",         (0,0),(-1,-1), 0.4, _hex_color("DDDDEE")),
        # Colorir % de progresso
        ("TEXTCOLOR",    (3,1),(3,1), ACCENT),
        ("FONTNAME",     (3,1),(3,1), "Helvetica-Bold"),
        ("FONTSIZE",     (3,1),(3,1), 13),
    ]))
    story.append(pt)
    story.append(Spacer(1, 12))

    # Distribuição por quadrante
    story.append(Paragraph("Distribuição por Quadrante",
                            ParagraphStyle("sh2", fontSize=11, fontName="Helvetica-Bold",
                                           textColor=DARK, spaceBefore=4, spaceAfter=6)))
    q_data = [["Quadrante", "Total", "Concluídas", "Tempo (h:mm)"]]
    for q in QUADRANT_ORDER:
        q_tasks = [t for t in tasks if str(t.quadrant).replace("EisenhowerQuadrant.", "") == q]
        q_done  = sum(1 for t in q_tasks if str(t.status) in ("done","TaskStatus.done"))
        q_time  = sum(t.time_spent_minutes for t in q_tasks)
        q_data.append([Q_LABELS[q], str(len(q_tasks)), str(q_done), _time_label(q_time)])

    qt = Table(q_data, colWidths=["50%","17%","17%","16%"])
    qt.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,0), _hex_color("4A4A7A")),
        ("TEXTCOLOR",    (0,0),(-1,0), WHITE),
        ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0),(-1,-1), 9),
        ("ALIGN",        (1,0),(-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",  (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [WHITE, LIGHT]),
        ("GRID",         (0,0),(-1,-1), 0.4, _hex_color("DDDDEE")),
    ]))
    story.append(qt)
    story.append(Spacer(1, 14))

    # Tasks em andamento
    if in_prog:
        story.append(Paragraph("Em Andamento",
                                ParagraphStyle("sh3", fontSize=11, fontName="Helvetica-Bold",
                                               textColor=DARK, spaceBefore=4, spaceAfter=6)))
        ip_data = [["Título", "Quadrante", "Prazo"]]
        for t in in_prog[:20]:  # máximo 20
            ip_data.append([
                Paragraph(t.title, ParagraphStyle("c", fontSize=8, fontName="Helvetica",
                                                   textColor=DARK, leading=10)),
                Q_LABELS.get(str(t.quadrant).replace("EisenhowerQuadrant.", ""), ""),
                t.due_date_iso or "—",
            ])
        ipt = Table(ip_data, colWidths=[9*cm, 5*cm, 3*cm])
        ipt.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,0), _hex_color("7C6DF0")),
            ("TEXTCOLOR",    (0,0),(-1,0), WHITE),
            ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0),(-1,0), 8),
            ("FONTSIZE",     (0,1),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [WHITE, LIGHT]),
            ("GRID",         (0,0),(-1,-1), 0.4, _hex_color("DDDDEE")),
            ("TOPPADDING",   (0,0),(-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ("LEFTPADDING",  (0,0),(-1,-1), 6),
        ]))
        story.append(ipt)

    # Rodapé
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", color=_hex_color("DDDDEE"), thickness=0.5))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"OrchFlow · {project.name} · {date.today().strftime('%d/%m/%Y')}",
        ParagraphStyle("ft", fontSize=8, fontName="Helvetica",
                       textColor=MUTED, alignment=1),
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ══ Endpoints ═════════════════════════════════════════════════════════════════

@router.get("/backlog/{project_id}")
def export_backlog(
    project_id: str,
    format: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    db: Session = Depends(get_db),
):
    """
    Exporta backlog completo do projeto em XLSX ou PDF.
    Somente leitura — zero escrita no banco.
    """
    project = _get_project(project_id, db)
    tasks   = _get_tasks(project_id, db)
    safe    = _safe_name(project.name)
    today   = date.today().isoformat()

    if format == "xlsx":
        buf = _build_backlog_xlsx(project, tasks)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="backlog_{safe}_{today}.xlsx"'},
        )
    else:
        buf = _build_backlog_pdf(project, tasks)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="backlog_{safe}_{today}.pdf"'},
        )


@router.get("/project/{project_id}/summary")
def export_project_summary(
    project_id: str,
    format: str = Query("pdf", pattern="^pdf$"),
    db: Session = Depends(get_db),
):
    """
    One-pager PDF do projeto: progresso, quadrantes, tasks em andamento.
    Somente leitura — zero escrita no banco.
    """
    project = _get_project(project_id, db)
    tasks   = _get_tasks(project_id, db)
    safe    = _safe_name(project.name)
    today   = date.today().isoformat()

    buf = _build_summary_pdf(project, tasks)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="summary_{safe}_{today}.pdf"'},
    )


@router.get("/time/{project_id}")
def export_time_report(
    project_id: str,
    format: str = Query("xlsx", pattern="^xlsx$"),
    db: Session = Depends(get_db),
):
    """
    Relatório de tempo por task em XLSX.
    Somente leitura — zero escrita no banco.
    """
    project = _get_project(project_id, db)
    tasks   = _get_tasks(project_id, db)
    safe    = _safe_name(project.name)
    today   = date.today().isoformat()

    buf = _build_time_xlsx(project, tasks)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="time_{safe}_{today}.xlsx"'},
    )
